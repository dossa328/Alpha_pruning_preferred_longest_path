import time
from copy import deepcopy as c
import numpy as np
import collections
import dijkstra
import openpyxl
count = 0
ret_path = []


def get_result(_metro, in_start, in_end, input_paths, trans_data, count, sheet):

    def cal_paths_cost(_metro, paths):
        sum = 0
        for kk in range(len(paths)-1):
            sum = sum + _metro.graph.get_cost(paths[kk], paths[kk+1])
        return sum

    def score_sub2(p_sub, _metro):
        cost = cal_path_weight2(p_sub, _metro)
        # cost2 = score_sub(p_sub)
        return cost

    def cal_path_weight2(_path, _metro):

        _trans_list = _metro.trans_list
        _graph = _metro.graph

        p_acc = 0.2 if _path[0] in _trans_list["trans"] else 0.5
        if len(_path) >= 2:
            expectation = p_acc * _graph.get_cost(_path[0], _path[1])
        elif len(_path) < 2:
            expectation = p_acc * 0

        for i in range(1, len(_path) - 1):
            p_i = 0.2 if _path[i] in _trans_list["trans"] else 0.5
            p_acc = p_acc + (1 - p_acc) * p_i
            expectation = expectation + p_acc * _graph.get_cost(_path[i], _path[i + 1])

        return expectation

    # split된 path 에 대해서 score를 계산한다.
    def score_sub(p_sub, _metro):
        cost = 0
        p_remaining = 1
        for i in range(len(p_sub)):
            p_i = 0.2 if p_sub[i] in _metro.trans_list["trans"] else 0.5
            p_i = p_remaining * p_i
            p_remaining = p_remaining - p_i

            cost = cost + p_i * cal_path_weight(i, p_sub)
        return cost

    def cal_path_weight(idx, path, _metro):
        sum_path = 0
        for i in range(idx, len(path) - 1):
            sum_path = sum_path + _metro.graph.get_cost(path[i], path[i + 1])
        return sum_path

    def split(path, _metro):
        out_cost = 0
        sv = []
        for i in range(len(path)):
            sv.append(path[i])
            # if i in trans_list["trans"] or i == in_end:
            if path[i] == in_end:
                out_cost = out_cost + score_sub2(sv, _metro)
                # paths.append(score_sub(sv))
                sv = []

            elif not _metro.station_to_line[path[i]] == _metro.station_to_line[path[i + 1]]:
                out_cost = out_cost + score_sub2(sv, _metro)
                # paths.append(score_sub(sv))
                sv = []
        return out_cost

    # python.org shortest_path advanced
    def find_shortest_path(graph2, start, end, _threshold, weight=0,  path=[[], 0]):
        dist = {start: [start]}
        q = collections.deque([start])
        while len(q):
            at = q.popleft()
            for next in graph2[at]:
                if next not in dist:
                    # dist[next] = [dist[at], next]
                    dist[next] = dist[at] + [next]
                    q.append(next)
        return dist[end]

    # shortest_path = find_shortest_path(_metro.graph.cost_matrix, in_start, in_end, 0)
    #
    # shortest_path_cost = cal_paths_cost(_metro, shortest_path)
    #
    # reresult = shortest_path
    #
    # index = 1
    # compare_paths_transdata = []
    # ways_2_transit_station = []
    # ways_3_transit_station = []
    # ways_4_transit_station = []
    # non_transdata_path = []
    # ways2 = ['서울A', '서울K', '동대문역사문화공원2', '동대문역사문화공원5', '김포공항A', '김포공항5', '까치산2', '까치산5', '효창공원앞6', '효창공원앞K', '충정로2', '충정로5', '을지로4가2', '을지로4가5', '청구5', '청구6', '신당2', '신당6', '합정2', '합정6', '영등포구청2', '영등포구청5']
    # ways3 = ['홍대입구2', '홍대입구A', '홍대입구K', '디지털미디어시티K', '디지털미디어시티A', '디지털미디어시티6', '왕십리2', '왕십리5', '왕십리K']
    # ways4 = ['공덕5', '공덕6', '공덕K', '공덕A']
    # for v in reresult:
    #     chk = 0
    #     for v2 in trans_data["trans"]:
    #         if v == in_end:
    #             pass
    #         else:
    #             if v2 in v:
    #                 if v2 in ways2:
    #                     ways_2_transit_station.append(v2)
    #                 elif v2 in ways3:
    #                     ways_3_transit_station.append(v2)
    #                 elif v2 in ways4:
    #                     ways_4_transit_station.append(v2)
    #                 chk = 1
    #                 compare_paths_transdata.append(v2)
    #
    #     if chk == 0:
    #         non_transdata_path.append(v)
    #     # print(compare_paths_transdata)
    #     # len_compare_paths_transdata = len(compare_paths_transdata)
    #
    # # wb = openpyxl.load_workbook('cls.xlsx')
    # # sheet = wb['Sheet1']
    count = count+2
    # for_return_values = []
    # # shp pairs
    # for_return_values.append([in_start, in_end])
    # sheet.cell(row=count, column=2, value=in_start)
    # sheet.cell(row=count, column=3, value=in_end)
    # # shp 역수
    # for_return_values.append(len(shortest_path))
    # sheet.cell(row=count, column=4, value=len(shortest_path))
    # # shp 경유 역수
    # for_return_values.append(len(compare_paths_transdata))
    # sheet.cell(row=count, column=5, value=len(compare_paths_transdata))
    # # shp 경유역 리스트
    # for_return_values.append(compare_paths_transdata)
    # sheet.cell(row=count, column=6, value=str(compare_paths_transdata))
    # # shp 혼잡경유 역수
    # # shp cost
    # for_return_values.append(shortest_path_cost)
    # sheet.cell(row=count, column=7, value=int(shortest_path_cost))
    # # shp paths
    # for_return_values.append(shortest_path)
    # sheet.cell(row=count, column=8, value=str(shortest_path))
    # # wb.save('cls.xlsx')
    # sheet.cell(row=count, column=12, value=len(ways_2_transit_station))
    # sheet.cell(row=count, column=13, value=len(ways_3_transit_station))
    # sheet.cell(row=count, column=14, value=len(ways_4_transit_station))
    #
    # print(for_return_values)

    # print(all_paths_candidate_paths)
    # print(len(all_paths_candidate_paths))
    # output = []
    shortest_path_0 = []
    # # shortest_path_0.append(input_paths)
    shortest_path_0 = input_paths.split()
    cal_paths_weight = cal_paths_cost(_metro, shortest_path_0)
    sheet.cell(row=count, column=18, value=cal_paths_weight)
    # for p in shortest_path_0:
    #     # saved.append(split(p[0]))
    #     path_cost = split(p, _metro)
    #     # p.append([(pow(p[1], -1)) * path_cost])
    #     p.append(path_cost)
    #     output.append(p)
    #
    # sheet.cell(row=count, column=16, value=path_cost)
    return 0
    # sort , 높은 점수만
    # output = sorted(output, key=lambda cp: -cp[2])
    # print(output)
    # return output[0]
