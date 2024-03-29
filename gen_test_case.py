import json
import random
import threading
import alpha_pruning
import cal_alpha_pruning
import cal_paths_weights
import cal_all_paths
import cal_shortest_path
import time
import cal_scores
import openpyxl
from Metro import Metro
import make_dataset


def rand_start_end(_data_set):
    while True:
        _start = random.choice(_data_set)
        _end = random.choice(_data_set)

        if _start is not _end:
            break

    return _start, _end


def _gen_test(fm_trans_json, fm_line_json, fm_edges_fix_json, input_vertex, turn_count):

    with open(fm_trans_json, 'r', encoding='UTF-8') as json_file:
        trans_data = json.load(json_file)

    with open(fm_line_json, 'r', encoding='UTF-8') as line_file:
        line_data = json.load(line_file)

    # with open('trans_classification_experiment_data.json', 'r', encoding='UTF-8') as data_file:
    # with open('trans_classification_list.json', 'r', encoding='UTF-8') as data_file:
    #     input_data = json.load(data_file)

    data_set = []

    for i in line_data:
        for j in line_data[i]:
            data_set.append(j + i)

    alpha_for_alpha_pruning = 1.2
    alpha_for_straight_forward = 1.2
    alpha_pruning_result = []
    alpha_pruning_result_avg = 0
    straight_forward_result = []
    straight_forward_result_avg = 0
    #
    # wb = openpyxl.load_workbook('result for cal.xlsx')
    # sheet = wb['Sheet1']

    # for k in range(1, 11):
    # start_time = time.time()
    reresult = []
    path_start_end_pairs = {}
    compare_paths_transdata = []


    a = 0
    # for j in range(0, len(input_data)):
    #     print(a)
        # reresult.extend(dijkstra_distributer.cal_dists(input_data[str(j)]['from'], input_data[str(j)]['to']))
        # reresult.extend(alpha_pruning.get_result(metro, input_data[str(j)]['from'], input_data[str(j)]['to'], alpha_for_alpha_pruning))
        # reresult.append(alpha_pruning2.get_result(metro, input_data[str(j)]['from'], input_data[str(j)]['to'], alpha_for_alpha_pruning)[0])
        # reresult.append(alpha_pruning2.get_result(metro, input_data[str(j)]['from'], input_data[str(j)]['to'], alpha_for_alpha_pruning)[0])
        # reresult.append(alpha_pruning2.get_result(metro, input_data[str(j)]['from'], input_data[str(j)]['to']))
        # a = a + 1
    # sett = []
    # for i in range(50000):
    #     ran_start, ran_end = rand_start_end(data_set)
    #     if not [ran_start, ran_end] in sett and not [ran_end, ran_start] in sett:
    #         sett.append([ran_start, ran_end])
    #         # print([ran_start, ran_end])
    #         print(len(sett))
    #     if len(sett) >= 10000:
    #         break

    # print(sett)
    # print(len(sett))

    # file_len = len(sett)
    file_len = 1999

    # make_dataset.make_data(len(sett), sett, input_vertex)

    metro = Metro(fm_trans_json, fm_line_json, fm_edges_fix_json)

    # fm_line_small_json = 'line_' + ','.join(input_vertex) + '_smalldata.json'
    #
    # with open(fm_line_small_json, 'r', encoding='UTF-8') as line_file:
    #     line_smalldata = json.load(line_file)

    ################ shortest_path구하는 곳
    # wb = openpyxl.load_workbook('cls.xlsx')
    # sheet = wb['Sheet1']
    # for i in range(file_len):
    #     # cal_shortest_path.get_result(metro, line_smalldata[str(i)]['from'], line_smalldata[str(i)]['to'], trans_data, i, sheet)
    #     cal_shortest_path.get_result(metro, sheet.cell(i + 2, 2).value, sheet.cell(i + 2, 3).value, trans_data, i, sheet)
    #     print(i)
    # wb.save('cls.xlsx')
    # /////////////////////////////////////
    # ############### cal_scores
    # wb = openpyxl.load_workbook('cls.xlsx')
    # sheet = wb['Sheet1']
    # for i in range(file_len):
    #     cal_scores.get_result(metro, sheet.cell(i + 2, 2).value, sheet.cell(i + 2, 3).value, sheet.cell(i+2, 8).value, trans_data, i, sheet)
    #     print(i)
    # wb.save('cls.xlsx')
    # /////////////////////////////////////
    # ############### cal_weights
    wb = openpyxl.load_workbook('cls.xlsx')
    sheet = wb['Sheet1']
    for i in range(file_len):
        cal_paths_weights.get_result(metro, sheet.cell(i + 2, 2).value, sheet.cell(i + 2, 3).value, sheet.cell(i+2, 15).value, trans_data, i, sheet)
        print(i)
    wb.save('cls.xlsx')
    # /////////////////////////////////////
    cal_time = 0
    cal_time2 = 0
    # 9734

    # wb = openpyxl.load_workbook('cls.xlsx')
    # sheet = wb['Sheet1']
    #
    # for i in range(file_len):
    #     # reresult.append(cal_all_paths.get_result(metro, line_smalldata[str(i)]['from'], line_smalldata[str(i)]['to'], alpha_for_straight_forward))
    #     # if sheet.cell(i+2, 5).value == 6 or sheet.cell(i+2, 5).value == 7 or sheet.cell(i+2, 5).value == 8 or sheet.cell(i+2, 5).value == 9 or sheet.cell(i+2, 5).value == 10:
    #     #     pass
    #     # else:
    #     if sheet.cell(i+2, 5).value == 9:
    #         # if sheet.cell(i+2, 9).value == 0:
    #         start_time = time.time()
    #         reresult.append(cal_all_paths.get_result(metro, sheet.cell(i + 2, 2).value, sheet.cell(i + 2, 3).value,
    #                                                  alpha_for_straight_forward))
    #         end_time = time.time()
    #         cal_time = end_time - start_time
    #         print(i, "th _all path finish", ":", sheet.cell(i + 2, 2).value, sheet.cell(i + 2, 3).value, "->",
    #               cal_time)
    #         sheet.cell(row=i + 2, column=9, value=cal_time)
    #         wb.save('cls.xlsx')
            # else:
            #     start_time = time.time()
            #     reresult.append(cal_all_paths.get_result(metro, sheet.cell(i+2, 2).value, sheet.cell(i+2, 3).value, alpha_for_straight_forward))
            #     end_time = time.time()
            #     cal_time = end_time - start_time
            #     print(i, "th _all path finish", ":", sheet.cell(i+2, 2).value, sheet.cell(i+2, 3).value, "->", cal_time)
            #     sheet.cell(row=i + 2, column=9, value=cal_time)
            #     wb.save('cls.xlsx')
                # if i % 2000 == 0:
                # print("하는중 -> ", i , len(sett))

    # ################ alpha
    # wb = openpyxl.load_workbook('cls.xlsx')
    # sheet = wb['Sheet1']
    # # for kkkk in range(100):
    # for j in range(file_len):
    #     # if sheet.cell(j+2, 5).value == 6 or sheet.cell(j+2, 5).value == 7 or sheet.cell(j+2, 5).value == 8 or sheet.cell(j+2, 5).value == 9 or sheet.cell(j+2, 5).value == 10:
    #     #     pass
    #     # else:
    #     if sheet.cell(j + 2, 5).value == 9:
    #     # if sheet.cell(j+2, 5).value == 0 or sheet.cell(j+2, 5).value == 1 or sheet.cell(j+2, 5).value == 2 or sheet.cell(j+2, 5).value == 3 or sheet.cell(j + 2, 5).value == 4 or sheet.cell(j+2, 5).value == 5 or sheet.cell(j+2, 5).value == 6 or sheet.cell(j+2, 5).value == 7 or sheet.cell(j+2, 5).value == 8:
    #         # if sheet.cell(j+2, 10).value == 0:
    #         start_time2 = time.time()
    #         reresult.append(
    #             cal_alpha_pruning.get_result(metro, sheet.cell(j + 2, 2).value, sheet.cell(j + 2, 3).value,
    #                                          alpha_for_alpha_pruning))
    #         end_time2 = time.time()
    #         cal_time2 = end_time2 - start_time2
    #         print(j, "th _alpha path finish", ":", sheet.cell(j + 2, 2).value, sheet.cell(j + 2, 3).value, "->",
    #               cal_time2)
    #         sheet.cell(row=j + 2, column=10, value=cal_time2)
    #         sheet.cell(row=j + 2, column=15, value=str(reresult[-1][0]))
    #
    # wb.save('cls.xlsx')
    # print("fin")
    # alpha /////////////////////////
        # reresult.append(cal_alpha_pruning.get_result(metro, line_smalldata[str(j)]['from'], line_smalldata[str(j)]['to'], alpha_for_alpha_pruning))
        #     else:
        #         start_time2 = time.time()
        #         reresult.append(cal_alpha_pruning.get_result(metro, sheet.cell(j + 2, 2).value, sheet.cell(j + 2, 3).value, alpha_for_alpha_pruning))
        #         end_time2 = time.time()
        #         cal_time2 = end_time2 - start_time2
        #         print(j, "th _alpha path finish", ":", sheet.cell(j+2, 2).value, sheet.cell(j+2, 3).value, "->", cal_time2)
        #         sheet.cell(row=j + 2, column=10, value=cal_time2)
        #         wb.save('cls.xlsx')
        #         # if i % 2000 == 0:
        #         # print("하는중 -> ", i , len(sett))


    # print(cal_time / file_len)
    # print(cal_time2 / file_len)
    # print(len(trans_data['trans']))
    # wb.save('cls.xlsx')

    # wb = openpyxl.load_workbook('smalldata_result.xlsx')
    # sheet = wb['Sheet4']
    # sheet.cell(row=turn_count, column=3, value=len(trans_data['trans']))
    # sheet.cell(row=turn_count, column=4, value=len(sett))
    # sheet.cell(row=turn_count, column=5, value=cal_time/file_len)
    # sheet.cell(row=turn_count, column=6, value=cal_time2/file_len)
    # wb.save('smalldata_result.xlsx')

# for i in reresult[0]:
#     print(i)
#
# print(len(reresult[0]))

# set_of_reresult = []
# for d in range(len(reresult)):
#     set_of_reresult.append(reresult[0])

#
# index = 1
# for v in reresult:
#     compare_paths_transdata = []
#     for v2 in trans_data["trans"]:
#         if v2 in v[0]:
#             compare_paths_transdata.append(v2)
#     print(compare_paths_transdata)
#     len_compare_paths_transdata = len(compare_paths_transdata)
#
#     path_start_end_pairs[len_compare_paths_transdata] = (v[0], v[-1])ㅔ
#     print(len_compare_paths_transdata, ",", v[0], ",", v[-1])
#     print(v[0][0], v[0][-1])
#     print(v)
#     # 지나가는 경유역 개수 (동일역 포함
#     sheet.cell(row=index, column=1, value=len_compare_paths_transdata)
#     sheet.cell(row=index, column=2, value=v[0][0])
#     sheet.cell(row=index, column=3, value=v[0][-1])
#     sheet.cell(row=index, column=4, value=v[-1])
#     sheet.cell(row=index, column=5, value=str(v[0]))
#
#
#     # sheet.cell(row=index, column=4, value=v)
#     # sheet.cell(row=index, column=4).value = v
#     # for r in range(0, len(v)):
#     #     sheet.cell(row=index, column=4).value=v[r]
#     index = index + 1
#
#     # alpha_pruning_result.append(alpha_pruning.get_result(metro, input_data[str(j)]['from'], input_data[str(j)]['to'], alpha_for_alpha_pruning))
#     # print(alpha_pruning.get_result(metro, input_data[str(j)]['from'], input_data[str(j)]['to'], alpha_for_alpha_pruning))
#     # alpha_pruning_result_avg = alpha_pruning_result_avg + alpha_pruning_result[j][2]
#
# wb.save('result for cal.xlsx')


# end_time = time.time()
# print(alpha_pruning_result_avg / len(input_data))
# print("AP_WorkingTime: {} sec".format(end_time-start_time))
    #
    # sheet.cell(row=k, column=1, value="알파 프루닝 결과 평균치")
    # sheet.cell(row=k, column=2, value=alpha_pruning_result_avg / len(input_data))
    # sheet.cell(row=k, column=3, value="알파 프루닝 결과 시간")
    # sheet.cell(row=k, column=4, value=format(end_time-start_time))
    #
    # wb.save('result.xlsx')

# start_time = time.time()
# for j in range(len(input_data)):
#     straight_forward_result.append(straight_forward.get_result(metro, input_data[str(j)]['from'], input_data[str(j)]['to'], alpha_for_straight_forward))
#     straight_forward_result_avg = straight_forward_result_avg + straight_forward_result[j][2]
#
# end_time = time.time()
# print(straight_forward_result_avg / len(input_data))
# print("SF_WorkingTime: {} sec".format(end_time - start_time))


# # 1.2 내 포함되는 데이터셋 만드는 코드
# time_out = 3
# deny = []
# for i in range(0, 10000):
#     inlist = []
#     ran_start, ran_end = rand_start_end(data_set)
#
#     done_counting = threading.Event()
#     # t = threading.Thread(target=alpha_pruning.get_result, args=(metro, ran_start, ran_end, alpha))
#     # t = threading.Thread(target=straight_forward.get_result, args=(metro, ran_start, ran_end, alpha_for_alpha_pruning))
#     t = threading.Thread(target=straight_forward_for_datasetting.get_result, args=(metro, ran_start, ran_end, alpha_for_alpha_pruning))
#     t.setDaemon(True)
#
#     t.start()
#
#     t.join(time_out)
#     done_counting.wait(time_out)
#     # if runtime out (difficult problem)
#     if t.is_alive():
#         # print(ran_start, ran_end)
#         pass
#     # if runtime out (easy problem)
#     else:
#         inlist = [ran_start, ran_end]
#         if not list in deny:
#             deny.append([ran_start, ran_end])
#             print(ran_start, ran_end)
#         else:
#             i = i - 1
#     # pass

# wb.save('result2.xlsx')